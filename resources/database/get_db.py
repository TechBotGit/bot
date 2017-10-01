import pandas as pd
import os
import sys
import json
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter, column_index_from_string

cwd = os.path.dirname(sys.argv[0])
os.chdir(cwd)

filename = "database_old.xlsx"

wb = load_workbook(filename)
sheet = wb.active
chat_id_test = 200158786
course_code_dict = {'MH1812': ['68ptpte11b7p7im3j84hnaddqs','3ifhcrdr18d4bmdkpsnafvltq4','houjv9j28ts33i16e1np54eafs, n2n389vqe4gr31n14i1i33nosk','27epdgmqfn864crq5s0n04h3so']}
course_code_dict_str = json.dumps(course_code_dict)
print("iter_cols()")
for col in sheet.iter_cols():
    for cell in col:
        print(cell.value)

print('---------------')
print("iter_rows()")
for row in sheet.iter_rows():
    for cell in row:
        if cell.value == chat_id_test:
            sheet.cell(row=cell.row, column=4).value = course_code_dict_str
            break

dictionary = sheet.cell(row=cell.row, column=4).value
dictionary = json.loads(dictionary)
print(len(dictionary))
wb.save(filename)
