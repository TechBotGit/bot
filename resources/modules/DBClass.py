import os
import sys
import json
from openpyxl import load_workbook, Workbook
# from openpyxl.utils.dataframe import dataframe_to_rows


class DB(object):

    def __init__(self):
        """Initialize WorkBook"""
        self.filename = "/../resources/database/database.xlsx"
        self.cwd = os.path.dirname(sys.argv[0])
        self.path_file = self.cwd + self.filename
        self.wb = Workbook()
        self.sheet = self.wb.active
        self.sheet['A1'] = 'chat_id'
        self.sheet['B1'] = 'first_week'
        self.sheet['C1'] = 'first_recess_week'
        self.sheet['D1'] = 'student_type'
        self.sheet['E1'] = 'course_code_event_id'
        self.sheet['F1'] = 'other_event_id'
        
        # If file doesn't exist, create it
        if not os.path.isfile(self.path_file):
            self.wb.save(self.path_file)
        
        # for updating method
        self.wb_update = load_workbook(self.path_file)
        self.sheet_update = self.wb_update.active

        self._chat_id_list = []

    @property
    def chat_id_list(self):
        return self._chat_id_list

    @chat_id_list.setter
    def chat_id_list(self, value):
        self._chat_id_list.append(value)
        return self._chat_id_list
    
    def update(self, chat_id, first_week=None, first_recess_week=None, student_type=None, course_code_event_id=None, other_event_id=None):
        """Description: Update exisiting workbook"""
        update_list = [chat_id, first_week, first_recess_week, student_type, course_code_event_id, other_event_id]
        if not self.isChatidExist(chat_id):
            self.sheet_update.append(update_list)
        else:
            print('Updating existing table')
            update_list.remove(chat_id)
            self.set_table_query(chat_id, update_list)

        self.wb_update.save(self.path_file)

    def isChatidExist(self, chat_id):
        for i in range(1, len(self.sheet_update['A'])):
            self.chat_id_list.append(self.sheet_update['A'][i].value)
        return chat_id in self.chat_id_list
    
    def isRecordExist(self, chat_id, first_week=None, first_recess_week=None, student_type=None, course_code_event_id=None, other_event_id=None):
        """Description: Check if a particular record exists in the database \n
        Usage: Set the optional parameter to be True to retrieve the data \n
        Return: Boolean
        Note: Only 1 optional parameter can be set to True at a time
        """
        result = None
        for row in self.sheet_update.iter_rows():
            for cell in row:
                if cell.value == chat_id:
                    if first_week:
                        result = self.sheet_update.cell(row=cell.row, column=2).value
                    elif first_recess_week:
                        result = self.sheet_update.cell(row=cell.row, column=3).value
                    elif student_type:
                        result = self.sheet_update.cell(row=cell.row, column=4).value
                    elif course_code_event_id:
                        dictionary = self.sheet_update.cell(row=cell.row, column=5).value
                        if dictionary != '{}' and dictionary is not None:
                            result = self.sheet_update.cell(row=cell.row, column=5).value
                    elif other_event_id:
                        dictionary = self.sheet_update.cell(row=cell.row, column=6).value
                        if dictionary != '{}' and dictionary is not None:
                            result = self.sheet_update.cell(row=cell.row, column=6).value
                    break
        return result is not None
    
    def table_query(self, chat_id, first_week=None, first_recess_week=None, student_type=None, course_code_event_id=None, other_event_id=None):
        """Description: Query table in database
        Usage: Set the requested data parameter to True to retrieve it.
        Return: list
        Note: Returns a list of requested data with the index coresponds to the order of the optional arguments, i.e. first_week has the index 0, first_recess_week has the index 1, etc."""
       
        arg_list = [first_week, first_recess_week, student_type, course_code_event_id, other_event_id]
        result_list = []
        for row in self.sheet_update.iter_rows():
            for cell in row:
                if cell.value == chat_id:
                    for i in range(len(arg_list)):
                        if arg_list[i] is not None:
                            result = self.sheet_update.cell(row=cell.row, column=i + 2).value
                            result_list.append(result)
                        else:
                            result_list.append(None)
                    break
                break
        return result_list
    
    def set_table_query(self, chat_id, update_list):
        """Description: Query table to set data with the corresponding chat_id \n
        Usage: Set the optional argument to the value that you want to set \n
        Example: set_table_query(<chat_id>, first_week='2017-8-14') \n
        Return: None
        """
        for row in self.sheet_update.iter_rows():
            for cell in row:
                if cell.value == chat_id:
                    for i in range(len(update_list)):
                        if update_list[i] is not None:
                            self.sheet_update.cell(row=cell.row, column=i + 2, value=update_list[i])
                    break
                break

    def UpdateCourseCodeEventId(self, chat_id, course_code, evt_id):
        if self.isChatidExist(chat_id):
            print('Updating existing table')
            for row in self.sheet_update.iter_rows():
                for cell in row:
                    if cell.value == chat_id:
                        data = self.sheet_update.cell(row=cell.row, column=5).value

                        # Parse to dictionary
                        data_dict = json.loads(data)
                        # Append the list inside the dictionary
                        data_dict[course_code]['event_id'].append(evt_id)
                        # Parse it back to strings
                        data_str = json.dumps(data_dict)
                        # Put it into the database
                        self.sheet_update.cell(row=cell.row, column=5, value=data_str)
                        break
                    break
        self.wb_update.save(self.path_file)
